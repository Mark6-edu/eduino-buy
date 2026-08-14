"""
Excel(xlsx) 구매 명세서 생성 기능을 담당하는 서비스 모듈.

추후 구현 예정:
- 학생별 구매 명세서 생성
- 기안문 스타일 품목내역 생성
- 교사용 전체 주문 통합 Excel 생성
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def create_order_excel(
    grade,
    class_name,
    student_number,
    student_name,
    selected_items,
    total_amount,
):
    """학생별 구매 명세서를 Excel 바이너리로 생성한다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "품목내역"

    header_font = Font(name="맑은 고딕", bold=True, size=12)
    title_font = Font(name="맑은 고딕", bold=True, size=16)
    body_font = Font(name="맑은 고딕", size=10)
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    fill = PatternFill("solid", fgColor="F2F2F2")

    sheet.sheet_view.showGridLines = False

    sheet.merge_cells("A1:G1")
    title_cell = sheet["A1"]
    title_cell.value = "Eduino 구매 명세서"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    info_rows = [
        ("학년", str(grade or "")),
        ("반", str(class_name or "")),
        ("번호", f"{student_number}번"),
        ("학생명", str(student_name or "")),
    ]

    for label, value in info_rows:
        sheet[f"A{row}"] = label
        sheet[f"B{row}"] = value
        sheet[f"A{row}"].font = body_font
        sheet[f"B{row}"].font = body_font
        row += 1

    merge_row = row + 1
    sheet.merge_cells(f"A{merge_row}:G{merge_row}")
    sheet[f"A{merge_row}"] = "<품목내역>"
    sheet[f"A{merge_row}"].font = header_font
    sheet[f"A{merge_row}"].alignment = Alignment(horizontal="center", vertical="center")

    header_row = merge_row + 2
    headers = [
        "순번",
        "내용",
        "규격",
        "수량",
        "단위",
        "예상단가",
        "예상금액",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = border

    current_row = header_row + 1
    for index, item in enumerate(selected_items, start=1):
        option_value = str(item.get("옵션", "")).strip()
        specification = item.get("상품코드") if (
            not option_value or option_value == "-"
        ) else option_value

        values = [
            f"1-{index}",
            str(item.get("상품명", "")),
            str(specification),
            int(item.get("수량", 0)),
            "개",
            int(item.get("단가", 0)),
            int(item.get("금액", 0)),
        ]

        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=current_row, column=col_index, value=value)
            cell.border = border
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")

            if col_index in (1, 4, 5):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_index == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_index in (6, 7):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

        sheet.cell(row=current_row, column=2).alignment = Alignment(
            wrap_text=True,
            vertical="center",
        )
        current_row += 1

    total_row = current_row + 1
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    total_label = sheet.cell(row=total_row, column=1, value="합계")
    total_label.font = header_font
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    total_label.border = border
    sheet.cell(row=total_row, column=7, value=int(total_amount or 0))
    sheet.cell(row=total_row, column=7,).font = header_font
    sheet.cell(row=total_row, column=7).number_format = "#,##0"
    sheet.cell(row=total_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    sheet.cell(row=total_row, column=7).border = border

    for row_idx in range(1, total_row + 1):
        sheet.row_dimensions[row_idx].height = 26

    column_widths = {
        "A": 10,
        "B": 45,
        "C": 20,
        "D": 10,
        "E": 10,
        "F": 15,
        "G": 15,
    }
    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    for row_idx in range(1, total_row + 1):
        for col_idx in range(1, 8):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.border = border

    sheet.freeze_panes = "A9"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
