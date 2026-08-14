import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st

from services.auth_service import is_teacher
from services.google_sheet_service import (
    fetch_order_summary,
    fetch_submissions,
    get_latest_submissions,
)
from services.excel_service import create_order_excel


if not is_teacher():
    st.warning("이 페이지는 교사 전용입니다.")
    st.stop()

st.set_page_config(
    page_title="교사용 주문 관리",
    page_icon="📊",
    layout="wide",
)


def _to_currency(value):
    return f"{int(value):,}원"


def format_quantity(value):
    return f"{int(value):,}개"


def format_short_datetime(value):
    if not value:
        return "-"

    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return str(value)[:16]


def display_student_label(student):
    grade = str(student.get("grade", ""))
    class_name = str(student.get("class_name", ""))
    student_number = str(student.get("student_number", ""))
    student_name = str(student.get("student_name", ""))
    return f"{grade} {class_name} {student_number}번 · {student_name}".strip()


def extract_numeric_sort_value(value):
    if value is None:
        return 0

    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip()
    if not text:
        return 0

    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else 0


@st.cache_data(ttl=30)
def load_all_order_data():
    submissions_result = fetch_submissions()
    summary_result = fetch_order_summary()

    if not submissions_result.get("success"):
        return {
            "success": False,
            "message": submissions_result.get("message", "학생 제출 데이터를 불러오지 못했습니다."),
            "submissions": [],
            "summary": [],
        }

    if not summary_result.get("success"):
        return {
            "success": False,
            "message": summary_result.get("message", "주문 집계 데이터를 불러오지 못했습니다."),
            "submissions": submissions_result.get("data", []),
            "summary": [],
        }

    submissions = get_latest_submissions(submissions_result.get("data", []))
    return {
        "success": True,
        "message": "OK",
        "submissions": submissions,
        "summary": summary_result.get("data", []),
    }


def _student_rows_for_display(rows):
    df = pd.DataFrame(rows)
    if df.empty:
        return df

    df["학년_num"] = df["grade"].str.extract(r"(\d+)", expand=False).fillna("0").astype(int)
    df["반_num"] = df["class_name"].str.extract(r"(\d+)", expand=False).fillna("0").astype(int)
    df["번호_num"] = df["student_number"].astype(str).str.extract(r"(\d+)", expand=False).fillna("0").astype(int)
    df = df.sort_values(["학년_num", "반_num", "번호_num"]).drop(columns=["학년_num", "반_num", "번호_num"])
    return df


def _aggregate_latest_submissions(rows):
    per_student = {}

    for row in rows:
        student_key = (
            row.get("grade", ""),
            row.get("class_name", ""),
            str(row.get("student_number", "")),
            row.get("student_name", ""),
            row.get("submission_id", ""),
        )

        entry = per_student.setdefault(
            student_key,
            {
                "grade": row.get("grade", ""),
                "class_name": row.get("class_name", ""),
                "student_number": str(row.get("student_number", "")),
                "student_name": row.get("student_name", ""),
                "submission_id": row.get("submission_id", ""),
                "submitted_at": row.get("submitted_at", ""),
                "items": [],
                "total_amount": 0,
            },
        )

        entry["items"].append(row)
        entry["total_amount"] += int(row.get("amount", 0) or 0)

    return list(per_student.values())


def _build_teacher_excel(latest_rows, summary_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "전체주문"

    header_font = Font(name="맑은 고딕", bold=True, size=11)
    title_font = Font(name="맑은 고딕", bold=True, size=16)
    body_font = Font(name="맑은 고딕", size=10)
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    fill = PatternFill("solid", fgColor="F2F2F2")

    summary_sheet.merge_cells("A1:H1")
    title_cell = summary_sheet["A1"]
    title_cell.value = "Eduino 전체 주문 통합"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["순번", "상품코드", "상품명", "옵션", "총수량", "단위", "예상단가", "예상금액"]
    start_row = 3
    for idx, header in enumerate(headers):
        cell = summary_sheet.cell(row=start_row, column=idx + 1, value=header)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_sum = 0
    for idx, row in enumerate(sorted(summary_rows, key=lambda item: (str(item.get("product_code", "")), str(item.get("option", "-")))), start=1):
        product_code = str(row.get("product_code", ""))
        product_name = str(row.get("product_name", ""))
        option = str(row.get("option", "-"))
        total_quantity = int(row.get("total_quantity", 0) or 0)
        unit_price = int(row.get("unit_price", 0) or 0)
        total_amount = int(row.get("total_amount", 0) or 0)
        total_sum += total_amount

        values = [idx, product_code, product_name, option, total_quantity, "개", unit_price, total_amount]
        for col_idx, value in enumerate(values, start=1):
            cell = summary_sheet.cell(row=start_row + idx, column=col_idx, value=value)
            cell.border = border
            cell.font = body_font
            if col_idx in (5, 7, 8):
                cell.number_format = "#,##0"
                if col_idx in (5, 7, 8):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    summary_end_row = start_row + len(summary_rows) + 1
    summary_sheet.merge_cells(start_row=summary_end_row, start_column=1, end_row=summary_end_row, end_column=7)
    summary_sheet.cell(row=summary_end_row, column=1, value="총 예상 구매금액").font = header_font
    summary_sheet.cell(row=summary_end_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    summary_sheet.cell(row=summary_end_row, column=8, value=total_sum).font = Font(name="맑은 고딕", bold=True, size=11)
    summary_sheet.cell(row=summary_end_row, column=8).number_format = "#,##0"
    summary_sheet.cell(row=summary_end_row, column=8).alignment = Alignment(horizontal="right", vertical="center")
    summary_sheet.row_dimensions[summary_end_row].height = 28

    widths = {"A": 10, "B": 18, "C": 32, "D": 18, "E": 12, "F": 10, "G": 16, "H": 16}
    for col, width in widths.items():
        summary_sheet.column_dimensions[col].width = width

    student_sheet = workbook.create_sheet("학생별제출")
    student_sheet.append(["학년", "반", "번호", "학생명", "제출일시", "상품코드", "상품명", "옵션", "수량", "단가", "금액"])

    sorted_rows = sorted(
        latest_rows,
        key=lambda row: (
            extract_numeric_sort_value(row.get("grade") or row.get("학년")),
            extract_numeric_sort_value(row.get("class_name") or row.get("반")),
            extract_numeric_sort_value(row.get("student_number") or row.get("번호")),
        ),
    )

    for row in sorted_rows:
        student_sheet.append([
            row.get("grade", ""),
            row.get("class_name", ""),
            row.get("student_number", ""),
            row.get("student_name", ""),
            row.get("submitted_at", ""),
            row.get("product_code", ""),
            row.get("product_name", ""),
            row.get("option", "-"),
            int(row.get("quantity", 0) or 0),
            int(row.get("unit_price", 0) or 0),
            int(row.get("amount", 0) or 0),
        ])

    for row in student_sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.font = body_font

    for cell in student_sheet[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    st_col_widths = {"A": 10, "B": 10, "C": 10, "D": 18, "E": 18, "F": 16, "G": 32, "H": 18, "I": 10, "J": 14, "K": 14}
    for col, width in st_col_widths.items():
        student_sheet.column_dimensions[col].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


# 초기 페이지 렌더링
st.title("👩‍🏫 교사용 주문관리")
st.caption("학생들이 제출한 Eduino 구매 내역을 확인하고 전체 주문 수량을 관리합니다.")

refresh_col, meta_col = st.columns([1.3, 4])
with refresh_col:
    if st.button("🔄 데이터 새로고침", use_container_width=True):
        st.cache_data.clear()
        st.rerun()
with meta_col:
    st.caption(f"마지막 조회: {format_short_datetime(datetime.now())}")

with st.spinner("제출 데이터를 불러오는 중입니다..."):
    data = load_all_order_data()

if not data.get("success"):
    st.warning(data.get("message", "학생 제출 데이터를 불러오지 못했습니다."))
    st.caption("잠시 후 다시 시도해주세요.")
    st.stop()

latest_rows = data.get("submissions", [])
summary_rows = data.get("summary", [])

student_rows = _aggregate_latest_submissions(latest_rows)

if not student_rows:
    st.info("아직 제출된 주문 내역이 없습니다.")
    st.caption("학생이 최종 제출하면 이 화면에 자동으로 표시됩니다.")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("제출 학생 수", "0명")
    col2.metric("주문 품목 수", "0종")
    col3.metric("총 주문 수량", "0개")
    col4.metric("전체 예상금액", "0원")
    st.stop()

# 필터
st.divider()
filter_cols = st.columns([1, 1, 1.4])
with filter_cols[0]:
    grade_filter = st.selectbox("학년", ["전체", "1학년", "2학년", "3학년", "기타"])
with filter_cols[1]:
    class_filter = st.selectbox("반", ["전체", "1반", "2반", "3반", "4반", "5반"])
with filter_cols[2]:
    name_filter = st.text_input("학생명 검색", placeholder="학생명 입력")

filtered_students = []
for student in student_rows:
    grade_ok = grade_filter == "전체" or student["grade"] == grade_filter
    class_ok = class_filter == "전체" or student["class_name"] == class_filter
    name_ok = not name_filter or name_filter in student["student_name"]
    if grade_ok and class_ok and name_ok:
        filtered_students.append(student)

filtered_students = sorted(
    filtered_students,
    key=lambda item: (
        int(str(item["grade"]).replace("학년", "") or 0),
        int(str(item["class_name"]).replace("반", "") or 0),
        int(str(item["student_number"]).replace("번", "") or 0),
    ),
)

summary_by_student_count = len(filtered_students)
summary_item_count = len({(item["product_code"], item["option"]) for student in filtered_students for item in student["items"]})
summary_total_quantity = sum(int(item["quantity"]) for student in filtered_students for item in student["items"])
summary_total_amount = sum(int(item["amount"]) for student in filtered_students for item in student["items"])

col1, col2, col3, col4 = st.columns(4)
col1.metric("제출 학생 수", f"{summary_by_student_count:,}명")
col2.metric("주문 품목 수", f"{summary_item_count:,}종")
col3.metric("총 주문 수량", f"{summary_total_quantity:,}개")
col4.metric("전체 예상금액", _to_currency(summary_total_amount))

st.divider()
st.subheader("👨‍🎓 학생 제출 현황")
st.caption("학생별 최신 제출 내역만 표시합니다.")

student_df = pd.DataFrame(
    [
        {
            "학년": student["grade"],
            "반": student["class_name"],
            "번호": student["student_number"],
            "학생명": student["student_name"],
            "제출일시": format_short_datetime(student["submitted_at"]),
            "품목수": len(student["items"]),
            "총수량": sum(int(item["quantity"]) for item in student["items"]),
            "총액": sum(int(item["amount"]) for item in student["items"]),
        }
        for student in filtered_students
    ]
)

if student_df.empty:
    st.info("조건에 맞는 학생 제출 내역이 없습니다.")
else:
    student_df["품목수"] = student_df["품목수"].map(lambda value: f"{int(value):,}종")
    student_df["총수량"] = student_df["총수량"].map(lambda value: f"{int(value):,}개")
    student_df["총액"] = student_df["총액"].map(lambda value: f"{int(value):,}원")
    st.dataframe(student_df, use_container_width=True, hide_index=True)

st.divider()
selected_student_label = None
if filtered_students:
    select_options = [display_student_label(student) for student in filtered_students]
    selected_student_label = st.selectbox(
        "학생별 주문 상세",
        options=select_options,
        index=0,
    )
    st.caption("조회할 학생을 선택하세요.")

if filtered_students and selected_student_label:
    selected = next(
        student for student in filtered_students if display_student_label(student) == selected_student_label
    )
    detail_df = pd.DataFrame(
        [
            {
                "상품코드": item["product_code"],
                "상품명": item["product_name"],
                "옵션": item["option"] if str(item.get("option", "")).strip() else "-",
                "단가": int(item["unit_price"]),
                "수량": int(item["quantity"]),
                "금액": int(item["amount"]),
            }
            for item in selected["items"]
        ]
    )
    st.subheader(f"🔎 {selected_student_label}")
    st.caption(f"제출일시: {format_short_datetime(selected.get('submitted_at'))}")
    detail_df["단가"] = detail_df["단가"].map(lambda value: f"{int(value):,}원")
    detail_df["수량"] = detail_df["수량"].map(lambda value: f"{int(value):,}개")
    detail_df["금액"] = detail_df["금액"].map(lambda value: f"{int(value):,}원")
    st.dataframe(detail_df, use_container_width=True, hide_index=True)
    student_total = sum(int(item["amount"]) for item in selected["items"])
    st.metric("학생 총액", f"{student_total:,}원")

st.divider()
st.subheader("📦 전체 주문 통합")
st.caption("최신 제출 기준으로 동일 상품과 옵션을 합산한 주문 목록입니다.")

summary_map = {}
for row in latest_rows:
    key = (str(row.get("product_code", "")), str(row.get("option", "-")))
    entry = summary_map.setdefault(
        key,
        {
            "product_code": str(row.get("product_code", "")),
            "product_name": str(row.get("product_name", "")),
            "option": str(row.get("option", "-")) if str(row.get("option", "")).strip() else "-",
            "unit_price": int(row.get("unit_price", 0) or 0),
            "total_quantity": 0,
            "total_amount": 0,
        },
    )
    entry["total_quantity"] += int(row.get("quantity", 0) or 0)
    entry["total_amount"] += int(row.get("amount", 0) or 0)

summary_table = pd.DataFrame(sorted(summary_map.values(), key=lambda x: (x["product_code"], x["option"])))
if summary_table.empty:
    st.info("아직 집계된 전체 주문 내역이 없습니다.")
else:
    summary_table = summary_table.rename(
        columns={
            "product_code": "상품코드",
            "product_name": "상품명",
            "option": "옵션",
            "unit_price": "예상단가",
            "total_quantity": "총수량",
            "total_amount": "예상금액",
        }
    )
    summary_table["예상단가"] = summary_table["예상단가"].map(lambda value: f"{int(value):,}원")
    summary_table["총수량"] = summary_table["총수량"].map(lambda value: f"{int(value):,}개")
    summary_table["예상금액"] = summary_table["예상금액"].map(lambda value: f"{int(value):,}원")
    st.dataframe(summary_table[["상품코드", "상품명", "옵션", "예상단가", "총수량", "예상금액"]], use_container_width=True, hide_index=True)
    total_order_amount = int(sum(item["total_amount"] for item in summary_map.values()))
    st.metric("전체 예상 구매금액", f"{total_order_amount:,}원")

excel_bytes = _build_teacher_excel(latest_rows, list(summary_map.values()))

st.divider()
st.subheader("📥 전체 주문 Excel 다운로드")
st.caption("통합 주문 내역과 학생별 최신 제출 내역이 하나의 Excel 파일에 포함됩니다.")
st.download_button(
    label="📗 전체 주문 Excel 다운로드",
    data=excel_bytes,
    file_name="Eduino_전체주문내역_2026-08-14.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    type="primary",
)
